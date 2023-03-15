import os
import sys
import re
import openpyxl
from docx import Document

# 检查并创建输出文件夹
output_dir = "output"
if not os.path.exists(output_dir):
    os.makedirs(output_dir)

# 打开输出文件
output_file_path = os.path.join(output_dir, "output.xlsx")
wb = openpyxl.Workbook()
ws = wb.active
ws.append(["姓名", "性别", "身份证号", "籍贯", "民族", "血型", "护理级别", "文化程度", "是否吸烟", "诊断"])

# 遍历当前目录中的Word文件
for file_name in os.listdir("."):
    if not file_name.endswith(".docx"):
        continue

    # 打开Word文件
    try:
        document = Document(file_name)
    except:
        print(f"无法打开文件 {file_name}")
        continue

    # 遍历Word文件中的所有表格
    for table in document.tables:
        # 查找需要的信息
        name = sex = id = jg = mz = blood = jb = whcd = cy = zd = None
        for row in table.rows:
            for i, cell in enumerate(row.cells):
                text = cell.text.strip()
                if text == "姓名":
                    name = row.cells[i+1].text.strip()
                elif text == "性别":
                    sex = row.cells[i+1].text.strip()
                elif text == "身份证号":
                    id = row.cells[i+1].text.strip()
                elif text == "籍贯":
                    jg = row.cells[i+1].text.strip()
                elif text == "民族":
                    mz = row.cells[i+1].text.strip()
                elif text == "血型":
                    blood = row.cells[i+1].text.strip()
                elif text == "护理级别":
                    jb = row.cells[i+1].text.strip()
                elif text == "文化程度":
                    whcd = row.cells[i+1].text.strip()
                elif text == "是否吸烟":
                    cy = row.cells[i+1].text.strip()
                elif text == "诊断":
                    zd = row.cells[i+1].text.strip()
                    zd = re.sub(r'\s+', ' ', zd)
                    

        # 输出结果到文件
        if name or zd:
            ws.append([name, sex, id, jg, mz, blood, jb, whcd, cy, zd])

    #关闭Word文件
    try:
        document.close()
    except:
        print(f"无法关闭文件 {file_name}")
        continue

# 保存.xlsx文件
for row in ws.iter_rows(min_row=3, min_col=10, max_col=10):
    if row[0].value:  # 如果I3单元格不为空
        # 将I3单元格的值复制到I2单元格
        ws.cell(row=row[0].row - 1, column=10).value = row[0].value
    # 删除第三行
    ws.delete_rows(row[0].row)
wb.save(output_file_path)

# 输出完成信息
print(f"输出完成，文件保存在：{output_file_path}")
